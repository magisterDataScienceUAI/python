from openpyxl import load_workbook;


import pandas as pd;
import datetime;

#df = pd.read_excel("Trabajo/metro.xlsx")

#print(df.iloc[[5,9]], 'pandas')


wb = load_workbook('metro..xlsx', data_only = True)

name = wb.sheetnames
print(name)

sheet = wb['Metro']
#print(sheet)

#meses=sheet.iloc[[5,19,1,1]].sum()

#print(meses)
#
rows=sheet.max_row
cols=sheet.max_column

var = 0

for i in range(1, cols):
  if isinstance(sheet.cell(row=4, column=i).value , str):
    print(sheet.cell(row=4, column=i).value)
    var1= i

metro = dict()

for i in range(4, rows):
  if isinstance(sheet.cell(row=i, column=2).value, datetime.datetime):
    print(sheet.cell(row=i, column=2).value)
    var2= i

for i in range(4,var2):
  for j in range(1,var1):
    metro[i,j] =  sheet.cell(row=i, column=j).value
  

df = pd.DataFrame.from_dict(metro, orient = 'index')
print(df)


    
